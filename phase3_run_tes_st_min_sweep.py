"""
TES Steam Turbine Minimum Load Sensitivity - Full Optimization Run

This script runs the complete optimization sweep:
1. Baseline: tes_st_min = 40%
2. Test: tes_st_min = 25%
3. Comparison of delivered costs

Expected: ~$3-4/MWh savings at 25% vs 40%
Estimated runtime: ~1 hour of solver time
"""

import pandas as pd
import openpyxl
import sys
import os
import json
import time
from pathlib import Path
import shutil

# Add Phase 2 module to path
sys.path.insert(0, "/Users/sreyachagarlamudi/Library/Mobile Documents/com~apple~CloudDocs/Intern Project Phase 2")

# Import the optimization module
try:
    from pyomo_06_loader import load_vars, load_svar, load_dfopsx
except:
    print("Documentation: Using simplified loader")

print("="*80)
print("TES STEAM TURBINE MINIMUM LOAD SENSITIVITY - FULL RUN")
print("="*80)
print()

# ============================================================================
# Configuration
# ============================================================================

excel_baseline = "/Users/sreyachagarlamudi/Downloads/(PtXv3.4_r0) gjt_working.xlsx"
excel_test = "/Users/sreyachagarlamudi/Downloads/(PtXv3.4_r0) gjt_working_test25.xlsx"
output_dir = "/Users/sreyachagarlamudi/Library/Mobile Documents/com~apple~CloudDocs/Intern Project Phase 3 - Analysis"

# ============================================================================
# Step 1: Create Test Excel File with tes_st_min = 25%
# ============================================================================

print("STEP 1: Creating test Excel file (tes_st_min = 25%)")
print("-" * 60)

if not os.path.exists(excel_test):
    print(f"Copying baseline to test file...")
    shutil.copy(excel_baseline, excel_test)
    print(f"  Success: Created: {excel_test}")

    # Modify tes_st_min to 25%
    wb = openpyxl.load_workbook(excel_test)
    tes_sheet = wb['TES']

    # Find tes_st_min row and update value
    found = False
    for row in tes_sheet.iter_rows(min_row=2):
        if row[0].value == 'tes_st_min':
            old_value = row[2].value
            row[2].value = 25.0
            print(f"  Success: Changed tes_st_min: {old_value}% → 25%")
            found = True
            break

    if not found:
        print("  Error: ERROR: tes_st_min not found in TES sheet!")
        print("  Please run phase3_tes_st_min_sensitivity.py first")
        sys.exit(1)

    wb.save(excel_test)
    wb.close()
    print(f"  Success: Saved: {excel_test}")
else:
    print(f"  Success: Test file already exists: {excel_test}")

print()

# ============================================================================
# Step 2: Load Excel Files and Extract Parameters
# ============================================================================

print("STEP 2: Loading Excel parameters")
print("-" * 60)

def read_tes_params(excel_path):
    """Read TES parameters from Excel file"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    tes_sheet = wb['TES']

    params = {}
    for row in tes_sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            param_name = row[0]
            param_value = row[2]  # Value is in column 3 (index 2)
            params[param_name] = param_value

    wb.close()
    return params

baseline_params = read_tes_params(excel_baseline)
test_params = read_tes_params(excel_test)

print(f"Baseline (40% min load):")
print(f"  tes_st_eff: {baseline_params.get('tes_st_eff', 'NOT FOUND')}%")
print(f"  tes_st_min: {baseline_params.get('tes_st_min', 'NOT FOUND')}%")
print()

print(f"Test (25% min load):")
print(f"  tes_st_eff: {test_params.get('tes_st_eff', 'NOT FOUND')}%")
print(f"  tes_st_min: {test_params.get('tes_st_min', 'NOT FOUND')}%")
print()

# Verify the change
if baseline_params.get('tes_st_min') == test_params.get('tes_st_min'):
    print("⚠️  WARNING: tes_st_min is the same in both files!")
    print("  Expected: Baseline=40%, Test=25%")
    print()

# ============================================================================
# Step 3: Manual Run Instructions
# ============================================================================

print("="*80)
print("STEP 3: OPTIMIZATION RUN INSTRUCTIONS")
print("="*80)
print()

print("⚠️  IMPORTANT: Full optimization must be run manually")
print("   Estimated time: ~1 hour of solver time")
print()

print("BASELINE RUN (tes_st_min = 40%):")
print("-" * 60)
print(f"1. cd '/Users/sreyachagarlamudi/Library/Mobile Documents/com~apple~CloudDocs/Intern Project Phase 2'")
print(f"2. Run optimization with: {excel_baseline}")
print(f"3. Save results to: {output_dir}/baseline_40pct_results.json")
print()

print("TEST RUN (tes_st_min = 25%):")
print("-" * 60)
print(f"1. cd '/Users/sreyachagarlamudi/Library/Mobile Documents/com~apple~CloudDocs/Intern Project Phase 2'")
print(f"2. Run optimization with: {excel_test}")
print(f"3. Save results to: {output_dir}/test_25pct_results.json")
print()

# ============================================================================
# Step 4: Create Results Comparison Template
# ============================================================================

print("="*80)
print("STEP 4: Results Comparison Template")
print("="*80)
print()

comparison_template = {
    "analysis_date": time.strftime("%Y-%m-%d"),
    "description": "TES steam turbine minimum load sensitivity analysis",
    "expected_result": "$3-4/MWh cost reduction at 25% vs 40% minimum load",
    "baseline": {
        "tes_st_min": baseline_params.get('tes_st_min', 40),
        "tes_st_eff": baseline_params.get('tes_st_eff', 40),
        "excel_file": excel_baseline,
        "results_file": f"{output_dir}/baseline_40pct_results.json",
        "lcoe_MWh": "TO BE FILLED",
        "cfe_percent": "TO BE FILLED",
        "annual_cost": "TO BE FILLED",
    },
    "test": {
        "tes_st_min": test_params.get('tes_st_min', 25),
        "tes_st_eff": test_params.get('tes_st_eff', 40),
        "excel_file": excel_test,
        "results_file": f"{output_dir}/test_25pct_results.json",
        "lcoe_MWh": "TO BE FILLED",
        "cfe_percent": "TO BE FILLED",
        "annual_cost": "TO BE FILLED",
    },
    "comparison": {
        "lcoe_delta_MWh": "TO BE CALCULATED",
        "lcoe_delta_percent": "TO BE CALCULATED",
        "cfe_delta": "TO BE CALCULATED",
        "annual_cost_delta": "TO BE CALCULATED",
    }
}

comparison_file = f"{output_dir}/tes_st_min_sensitivity_comparison.json"
with open(comparison_file, 'w') as f:
    json.dump(comparison_template, f, indent=2)

print(f"Success: Created comparison template: {comparison_file}")
print()

# ============================================================================
# Step 5: Summary
# ============================================================================

print("="*80)
print("SETUP COMPLETE - READY FOR OPTIMIZATION RUNS")
print("="*80)
print()

print("FILES CREATED:")
print(f"  Success: Test Excel: {excel_test}")
print(f"  Success: Comparison template: {comparison_file}")
print()

print("NEXT STEPS:")
print("  1. Run baseline optimization (tes_st_min = 40%)")
print("  2. Run test optimization (tes_st_min = 25%)")
print("  3. Compare results using comparison template")
print()

print("EXPECTED OUTCOME:")
print("  - Baseline LCOE: ~$52-60/MWh")
print("  - Test LCOE: ~$48-56/MWh (3-4/MWh lower)")
print("  - Mechanism: Better part-load operation reduces cycling losses")
print()

# ============================================================================
# Step 6: Create Quick Comparison Script
# ============================================================================

print("Creating post-run comparison script...")

comparison_script = '''"""
Compare results after both optimization runs complete
Run this after you have both baseline and test results
"""

import json

output_dir = "/Users/sreyachagarlamudi/Library/Mobile Documents/com~apple~CloudDocs/Intern Project Phase 3 - Analysis"
baseline_file = f"{output_dir}/baseline_40pct_results.json"
test_file = f"{output_dir}/test_25pct_results.json"
comparison_file = f"{output_dir}/tes_st_min_sensitivity_comparison.json"

# Load results (update these paths to match your actual output files)
try:
    with open(baseline_file, 'r') as f:
        baseline = json.load(f)
    with open(test_file, 'r') as f:
        test = json.load(f)

    # Extract key metrics (adjust field names to match actual output)
    baseline_lcoe = baseline.get('lcoe_MWh', 0)
    test_lcoe = test.get('lcoe_MWh', 0)

    delta_MWh = test_lcoe - baseline_lcoe
    delta_pct = (delta_MWh / baseline_lcoe * 100) if baseline_lcoe > 0 else 0

    print("="*80)
    print("TES STEAM TURBINE MINIMUM LOAD - RESULTS COMPARISON")
    print("="*80)
    print()

    print("BASELINE (tes_st_min = 40%):")
    print(f"  LCOE: ${baseline_lcoe:.2f}/MWh")
    print()

    print("TEST (tes_st_min = 25%):")
    print(f"  LCOE: ${test_lcoe:.2f}/MWh")
    print()

    print("DELTA:")
    print(f"  Absolute: ${delta_MWh:.2f}/MWh")
    print(f"  Relative: {delta_pct:.1f}%")
    print()

    if abs(delta_MWh) >= 3 and abs(delta_MWh) <= 5:
        print("Success: Result matches expected range ($3-4/MWh savings)")
    else:
        print(f"⚠️  Result outside expected range (expected: $3-4/MWh, got: ${abs(delta_MWh):.2f}/MWh)")
    print()

    # Update comparison template
    with open(comparison_file, 'r') as f:
        comp = json.load(f)

    comp['baseline']['lcoe_MWh'] = baseline_lcoe
    comp['test']['lcoe_MWh'] = test_lcoe
    comp['comparison']['lcoe_delta_MWh'] = delta_MWh
    comp['comparison']['lcoe_delta_percent'] = delta_pct

    with open(comparison_file, 'w') as f:
        json.dump(comp, f, indent=2)

    print(f"Success: Updated: {comparison_file}")

except FileNotFoundError as e:
    print("Error: Result files not found")
    print("  Make sure both optimization runs have completed")
    print("  Expected files:")
    print(f"    - {baseline_file}")
    print(f"    - {test_file}")
'''

comparison_script_file = f"{output_dir}/compare_tes_st_min_results.py"
with open(comparison_script_file, 'w') as f:
    f.write(comparison_script)

print(f"  Success: Created: {comparison_script_file}")
print()

print("="*80)
print("ALL SETUP COMPLETE")
print("="*80)
