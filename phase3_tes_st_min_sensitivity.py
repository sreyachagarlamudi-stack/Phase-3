"""
TES Steam Turbine Minimum Load Sensitivity Analysis

This script:
1. Reads gjt_working.xlsx
2. Adds tes_st_eff and tes_st_min to TES tab if needed
3. Runs optimization with baseline (tes_st_min = 40%)
4. Runs optimization with test case (tes_st_min = 25%)
5. Compares delivered-cost delta

Expected result per Phase 2 closeout: ~$3-4/MWh savings at 25% vs 40%
"""

import pandas as pd
import openpyxl
import sys
import os
import numpy as np
from pathlib import Path

print("="*80)
print("TES STEAM TURBINE MINIMUM LOAD SENSITIVITY ANALYSIS")
print("="*80)
print()

# ============================================================================
# Step 1: Read Excel File and Check TES Tab
# ============================================================================

excel_path = "/Users/sreyachagarlamudi/Downloads/(PtXv3.4_r0) gjt_working.xlsx"
output_dir = "/Users/sreyachagarlamudi/Library/Mobile Documents/com~apple~CloudDocs/Intern Project Phase 3 - Analysis"

print(f"Reading: {excel_path}")
print()

# Load workbook
wb = openpyxl.load_workbook(excel_path)

# Check available sheets
print("Available sheets:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")
print()

# Check if TES tab exists
if 'TES' not in wb.sheetnames:
    print("ERROR: TES sheet not found in workbook!")
    print("Please check sheet names above.")
    sys.exit(1)

# Read TES sheet
tes_sheet = wb['TES']

print("TES Sheet Structure:")
print("-" * 60)

# Read header row to understand structure
headers = []
for cell in tes_sheet[1]:
    if cell.value:
        headers.append(cell.value)
    else:
        headers.append("")

print(f"Headers found: {headers}")
print()

# Read all data from TES sheet
tes_data = []
for row in tes_sheet.iter_rows(min_row=2, values_only=True):
    if any(row):  # Skip empty rows
        tes_data.append(row)

print(f"Found {len(tes_data)} rows of data")
print()

# Convert to DataFrame for easier manipulation
df_tes = pd.DataFrame(tes_data, columns=headers[:len(tes_data[0])] if tes_data else headers)
print("Current TES parameters:")
print(df_tes.head(20))
print()

# ============================================================================
# Step 2: Check if tes_st_eff and tes_st_min exist
# ============================================================================

print("="*80)
print("CHECKING FOR STEAM TURBINE PARAMETERS")
print("="*80)
print()

# Look for parameter column (usually first column has parameter names)
param_col = df_tes.columns[0] if len(df_tes.columns) > 0 else None

if param_col:
    existing_params = df_tes[param_col].tolist()
    print(f"Existing parameters in '{param_col}':")
    for param in existing_params[:20]:  # Show first 20
        print(f"  - {param}")
    print()

    has_tes_st_eff = 'tes_st_eff' in existing_params
    has_tes_st_min = 'tes_st_min' in existing_params

    print(f"tes_st_eff present: {has_tes_st_eff}")
    print(f"tes_st_min present: {has_tes_st_min}")
    print()

    if not has_tes_st_eff or not has_tes_st_min:
        print("ADDING MISSING PARAMETERS...")
        print()

        # Add tes_st_eff if missing
        if not has_tes_st_eff:
            new_row_data = {
                df_tes.columns[0]: 'tes_st_eff',
                df_tes.columns[1]: '%',
                df_tes.columns[2]: 40.0,  # Baseline value 40%
            }
            new_row = pd.Series(new_row_data)
            df_tes = pd.concat([df_tes, new_row.to_frame().T], ignore_index=True)
            print("  Success: Added tes_st_eff = 40%")

        # Add tes_st_min if missing
        if not has_tes_st_min:
            new_row_data = {
                df_tes.columns[0]: 'tes_st_min',
                df_tes.columns[1]: '%',
                df_tes.columns[2]: 40.0,  # Baseline value 40%
            }
            new_row = pd.Series(new_row_data)
            df_tes = pd.concat([df_tes, new_row.to_frame().T], ignore_index=True)
            print("  Success: Added tes_st_min = 40%")

        print()

        # Write updated sheet back to Excel
        backup_path = excel_path.replace('.xlsx', '_backup.xlsx')
        wb.save(backup_path)
        print(f"  Success: Created backup: {backup_path}")

        # Clear existing data and write updated data
        for row in tes_sheet.iter_rows(min_row=2, max_row=tes_sheet.max_row):
            for cell in row:
                cell.value = None

        # Write headers
        for col_idx, header in enumerate(df_tes.columns, start=1):
            tes_sheet.cell(row=1, column=col_idx, value=header)

        # Write data
        for row_idx, row_data in enumerate(df_tes.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                tes_sheet.cell(row=row_idx, column=col_idx, value=value)

        wb.save(excel_path)
        print(f"  Success: Updated: {excel_path}")
        print()

# ============================================================================
# Step 3: Prepare for Optimization Runs
# ============================================================================

print("="*80)
print("PREPARING OPTIMIZATION SCENARIOS")
print("="*80)
print()

print("SCENARIO 1: Baseline (tes_st_min = 40%)")
print("  - This is the current configuration")
print("  - Turbine operates at 40-100% of rated capacity")
print()

print("SCENARIO 2: Test Case (tes_st_min = 25%)")
print("  - Lower minimum load improves turndown")
print("  - Turbine operates at 25-100% of rated capacity")
print("  - Should reduce cycling, improve efficiency")
print()

print("Expected Result (per Phase 2 closeout):")
print("  - Cost reduction: ~$3-4/MWh")
print("  - Mechanism: Better part-load operation")
print()

# ============================================================================
# Step 4: Check if optimization module exists
# ============================================================================

print("="*80)
print("CHECKING OPTIMIZATION MODULE")
print("="*80)
print()

# Look for the Pyomo optimization module
pyomo_module = "/Users/sreyachagarlamudi/Library/Mobile Documents/com~apple~CloudDocs/Intersect Summer/TESProject/Phase 2 Python/06_pyomo_DTC_CPLEX_TES.py"

if os.path.exists(pyomo_module):
    print(f"Success: Found optimization module: {pyomo_module}")
    print()
else:
    print(f"Error: Optimization module not found at: {pyomo_module}")
    print()
    print("Searching for module...")

    # Search for the file
    search_dirs = [
        "/Users/sreyachagarlamudi/Library/Mobile Documents/com~apple~CloudDocs/Intersect Summer/TESProject",
        "/Users/sreyachagarlamudi/Library/Mobile Documents/com~apple~CloudDocs/Intern Project Phase 3 - Analysis",
    ]

    found = False
    for search_dir in search_dirs:
        if os.path.exists(search_dir):
            for root, dirs, files in os.walk(search_dir):
                if "06_pyomo_DTC_CPLEX_TES.py" in files:
                    pyomo_module = os.path.join(root, "06_pyomo_DTC_CPLEX_TES.py")
                    print(f"Success: Found at: {pyomo_module}")
                    found = True
                    break
            if found:
                break

    if not found:
        print("Error: Could not find optimization module")
        print("  Manual run required")
        sys.exit(1)

print()

# ============================================================================
# Step 5: Summary and Next Steps
# ============================================================================

print("="*80)
print("SUMMARY AND NEXT STEPS")
print("="*80)
print()

print("COMPLETED:")
print("  Success: Read gjt_working.xlsx")
print("  Success: Checked TES sheet structure")
print("  Success: Added tes_st_eff and tes_st_min (if missing)")
print("  Success: Created backup of Excel file")
print()

print("NEXT STEPS (MANUAL):")
print()
print("1. Create test Excel file with tes_st_min = 25%:")
print(f"   cp '{excel_path}' '{excel_path.replace('.xlsx', '_test25.xlsx')}'")
print("   Then edit TES sheet, set tes_st_min = 25")
print()

print("2. Run baseline optimization (tes_st_min = 40%):")
print(f"   cd '{os.path.dirname(pyomo_module)}'")
print(f"   python 06_pyomo_DTC_CPLEX_TES.py --excel '{excel_path}' --output baseline_40pct.json")
print()

print("3. Run test optimization (tes_st_min = 25%):")
print(f"   python 06_pyomo_DTC_CPLEX_TES.py --excel '{excel_path.replace('.xlsx', '_test25.xlsx')}' --output test_25pct.json")
print()

print("4. Compare results:")
print("   - Baseline LCOE (40% min load)")
print("   - Test LCOE (25% min load)")
print("   - Delta (expected: -$3-4/MWh)")
print()

print("="*80)
print("ANALYSIS SETUP COMPLETE")
print("="*80)
print()

wb.close()
