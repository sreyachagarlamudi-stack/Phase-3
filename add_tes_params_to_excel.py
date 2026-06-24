"""
Add TES parameters sheet to gjt_working Excel file
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Load the Excel file
excel_path = "/Users/sreyachagarlamudi/Library/Mobile Documents/com~apple~CloudDocs/Intern Project Phase 3 - Analysis/gjt_working_updated.xlsx"

# Create TES parameters dataframe
tes_params = pd.DataFrame({
    'Parameter': [
        'tes_heater_capex',
        'tes_heater_opex_fixed',
        'tes_heater_efficiency',
        'tes_turbine_capex',
        'tes_turbine_opex_fixed',
        'tes_turbine_opex_var',
        'tes_turbine_size_kW',
        'tes_turbine_eff_min',
        'tes_turbine_eff_max',
        'tes_turbine_min_load',
        'tes_storage_capex',
        'tes_storage_opex_pct',
        'tes_storage_loss_per_day',
        'boiler_capex',
        'boiler_efficiency',
        'boiler_opex_fixed',
        'boiler_opex_var',
        'boiler_fuel_cfe',
    ],
    'Value': [
        200,      # Heater CapEx $/kW
        5,        # Heater OpEx fixed $/kW-year
        0.95,     # Heater efficiency
        800,      # Turbine CapEx $/kW
        15,       # Turbine OpEx fixed $/kW-year
        3,        # Turbine OpEx variable $/MWh
        100000,   # Turbine size kW (fixed at load)
        0.34,     # Turbine efficiency at 40% load
        0.39,     # Turbine efficiency at 100% load
        0.40,     # Turbine minimum load fraction
        40,       # Storage CapEx $/kWh
        0.005,    # Storage OpEx (0.5% per year)
        0.01,     # Storage heat loss (1% per day)
        300,      # Boiler CapEx $/kW_thermal
        0.85,     # Boiler efficiency
        10,       # Boiler OpEx fixed $/kW-year
        2,        # Boiler OpEx variable $/MWh
        0.0,      # Boiler fuel CFE (0 for NG)
    ],
    'Unit': [
        '$/kW',
        '$/kW-year',
        'fraction',
        '$/kW',
        '$/kW-year',
        '$/MWh',
        'kW',
        'fraction',
        'fraction',
        'fraction',
        '$/kWh',
        'fraction',
        'fraction',
        '$/kW_th',
        'fraction',
        '$/kW-year',
        '$/MWh_th',
        'fraction',
    ],
    'Source': [
        'Industry estimate - resistive heaters',
        'Minimal maintenance',
        'Standard resistive heating',
        'Similar to gas turbine',
        'Similar to gas turbine',
        'Maintenance and wear',
        'Fixed at datacenter load',
        'NREL/EPA at 40% load',
        'NREL/EPA at 100% load',
        'Physics constraint - steam turbulence',
        'Thermal storage media',
        'Minimal maintenance',
        'Insulation heat loss',
        'Much cheaper than NGPP',
        'NG to thermal conversion',
        'Boiler maintenance',
        'Boiler variable costs',
        '0 for natural gas, adjust for renewable diesel',
    ]
})

# Load workbook and add new sheet
wb = load_workbook(excel_path)

# Remove TES_PARAMS sheet if it exists
if 'TES_PARAMS' in wb.sheetnames:
    del wb['TES_PARAMS']

# Create new sheet
ws = wb.create_sheet('TES_PARAMS', 0)  # Add as first sheet

# Write headers
ws['A1'] = 'Parameter'
ws['B1'] = 'Value'
ws['C1'] = 'Unit'
ws['D1'] = 'Source'

# Write data
for idx, row in tes_params.iterrows():
    ws[f'A{idx+2}'] = row['Parameter']
    ws[f'B{idx+2}'] = row['Value']
    ws[f'C{idx+2}'] = row['Unit']
    ws[f'D{idx+2}'] = row['Source']

# Save workbook
wb.save(excel_path)
print(f"✓ Added TES_PARAMS sheet to {excel_path}")

# Create C4.csv for gas pricing ($4/MMBTU)
gas_price = pd.DataFrame({
    'hour': range(8760),
    'price': [4.0] * 8760
})
gas_price_path = "/Users/sreyachagarlamudi/Library/Mobile Documents/com~apple~CloudDocs/Intern Project Phase 3 - Analysis/C4.csv"
gas_price.to_csv(gas_price_path, index=False, header=False)
print(f"✓ Created {gas_price_path} with $4/MMBTU gas pricing")

print("\n✓ Excel file updated successfully!")
print("✓ Gas pricing file created!")
